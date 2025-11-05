import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Number of iterations
n_iter = 1000

# Function to get column letter
def get_col_letter(col_num):
    """Convert column number to Excel column letter"""
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result

def create_sheet(wb, sheet_name, sheet_config):
    """Create a sheet with specific configuration"""
    
    # Create or get worksheet
    if sheet_name == "NTS-L":
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    # Set up the exact structure as Sukuk_NTS.xlsx
    # Row 1: Title
    ws['H1'] = "RENTAL TAX SHIELD WITH LEVERAGE"
    
    # Row 2: Empty
    
    # Row 3: Column headers (A1, A2, A3, etc.)
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)  # Start from column I (9th column)
        ws[f'{col_letter}3'] = f'A{i+1}'
    
    # Row 4: Net Operating Income
    ws['A4'] = "Net Operating Income"
    ws['H4'] = "NOI"
    # Fill NOI values: linear increase from 10000 to 52000
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        value = 10000 + (52000 - 10000) * i / (n_iter - 1)
        ws[f'{col_letter}4'] = value
    
    # Row 5: Interest of Debt
    ws['A5'] = "Interest of Debt (market value of debt*interest rate)"
    ws['G5'] = "INTEREST ON DEBT"
    ws['H5'] = "INTEREST"
    # Formula: =I19*I24 (and similar for other columns)
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}5'] = f'={col_letter}19*{col_letter}24'
    
    # Row 6: Rent of Sukuk
    ws['A6'] = "Rent of Sukuk ijarah assets (market value of sukuk*rent rate)"
    ws['G6'] = "RENT ON SUKUK"
    ws['H6'] = "RENT"
    # Formula: =I20*I25 (and similar for other columns) - but some sheets use I31
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        if sheet_config.get('rent_uses_31', False):
            ws[f'{col_letter}6'] = f'={col_letter}20*{col_letter}31'
        else:
            ws[f'{col_letter}6'] = f'={col_letter}20*{col_letter}25'
    
    # Row 7: Benefit of asset depreciation
    ws['A7'] = "Benefit of asset depriciation/running expensed of the ijarah asset"
    ws['H7'] = "NDTS"
    # Formula: =I20*I26 (market value of sukuk * depreciation rate)
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}7'] = f'={col_letter}20*{col_letter}26'
    
    # Row 8: CAPITAL GAIN OR LOSS
    ws['A8'] = "CAPITAL GAIN OR LOSS"
    ws['G8'] = "CAPITAL GAIN OR LOSS"
    ws['H8'] = "CAPITAL GAIN OR LOSS"
    # Formula: =I20*I27 (market value of sukuk * capital gain/loss rate)
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}8'] = f'={col_letter}20*{col_letter}27'
    
    # Row 9: Dividend paid - THIS IS THE KEY DIFFERENCE BETWEEN SHEETS
    ws['A9'] = "Dividend paid (tax shield)"
    ws['G9'] = "DIVIDEND ON TAX"
    ws['H9'] = "Dividend"
    
    # Apply specific dividend formula based on sheet type
    if sheet_config.get('no_dividend_before_tax', False):
        # RTS-L and RTS-ZL: No dividend calculated before tax
        pass  # Leave empty
    elif sheet_config.get('dts_dividend_formula', False):
        # DTS-L and DTS-ZL: Special dividend formula
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}9'] = f'=({col_letter}4-{col_letter}7+{col_letter}8-{col_letter}13-{col_letter}14)*{col_letter}28'
    elif sheet_config.get('dts_rts_dividend_formula', False):
        # (DTS+RTS)-L and (DTS+RTS)-ZL: Combined formula
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}9'] = f'=({col_letter}4-{col_letter}6-{col_letter}7+{col_letter}8-{col_letter}13)*{col_letter}29'
    else:
        # Standard formula for NTS-L, ITS-L, NTS-ZL, ITS-ZL
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}9'] = f'=({col_letter}4-{col_letter}5-{col_letter}6-{col_letter}7)*{col_letter}28'
    
    # Row 10: Earnings before Tax
    ws['A10'] = "Earnings before Tax with above calculation"
    ws['H10'] = "EBT"
    # Formula: =I4-I5-I6-I7+I8-I9 (and similar for other columns)
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}10'] = f'={col_letter}4-{col_letter}5-{col_letter}6-{col_letter}7+{col_letter}8-{col_letter}9'
    
    # Row 11: Tax Amount
    ws['A11'] = f"TAX.C-{sheet_name}"
    ws['G11'] = "TAX AMOUNT Tax @ 35%"
    ws['H11'] = f"TAX.C-{sheet_name}"
    # Formula: =I10*I22 (EBT * Tax rate)
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}11'] = f'={col_letter}10*{col_letter}22'
    
    # Row 12: EAT
    ws['A12'] = "EAT"
    ws['H12'] = "EAT"
    # Formula: =I10-I11 (EBT - Tax Amount)
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}12'] = f'={col_letter}10-{col_letter}11'
    
    # Row 13: EARNING AVAILABLE FOR DEBTHOLDERS
    ws['A13'] = "EARNING AVAILABLE FOR DEBTHOLDERS"
    ws['G13'] = "EARNING AVAILABLE FOR DEBTHOLDERS"
    ws['H13'] = "INTEREST IF AFTER TAX"
    
    # Different formulas based on sheet type
    if sheet_config.get('interest_before_tax', False):
        # ITS-L and ITS-ZL: No formula (interest already deducted before tax)
        pass  # Leave empty
    else:
        # All other sheets: =I19*I30 (since I30=E24=0.1 always)
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}13'] = f'={col_letter}19*{col_letter}30'
    
    # Row 14: EARNING AVAILABLE FOR SUKUK HOLDERS
    ws['A14'] = "EARNING AVAILABLE FOR SUKUK HOLDERS"
    ws['G14'] = "EARNING AVAILABLE FOR SUKUK HOLDERS"
    ws['H14'] = "RENT IF AFTER TAX"
    
    # For RTS-L and RTS-ZL: Row 14 should be blank
    if sheet_name not in ['RTS-L', 'RTS-ZL']:
        # Formula: =I31*I20 (ks -sheet_name * Market value of sukuk)
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}14'] = f'={col_letter}31*{col_letter}20'
    # For RTS-L and RTS-ZL, leave Row 14 blank (empty cells)
    
    # Row 15: EARNING AVAILABLE FOR SHAREHOLDERS
    ws['A15'] = "EARNING AVAILABLE FOR SHAREHOLDERS"
    ws['G15'] = "EARNING AVAILABLE FOR SHAREHOLDERS"
    ws['H15'] = "DIVIDEND IF AFTER TAX"
    
    # Different formulas based on sheet type
    if sheet_name == 'DTS-ZL':
        # For DTS-ZL only: I15 = I12-I13-I14+I9, J15 = J12-J13-J14+J9, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}15'] = f'={col_letter}12-{col_letter}13-{col_letter}14+{col_letter}9'
    elif sheet_name in ['(DTS+RTS)-L', '(DTS+RTS)-ZL']:
        # For (DTS+RTS)-L and (DTS+RTS)-ZL: Row 15 should be blank (empty)
        pass  # Leave Row 15 blank (no formulas)
    else:
        # Formula: =I12-I13-I14 (EAT - Interest after tax - Rent after tax)
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}15'] = f'={col_letter}12-{col_letter}13-{col_letter}14'
    
    # Row 16: Dividend PAID
    ws['A16'] = "Dividend PAID"
    ws['G16'] = "Dividend PAID"
    ws['H16'] = "DIVIDEND IF AFTER TAX"
    
    # Different dividend paid formulas based on sheet type
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        
        if sheet_config.get('dts_dividend_formula', False) and not sheet_config.get('dts_rts_dividend_formula', False):
            # DTS-L and DTS-ZL: =I15+I9 (Earnings + dividend before tax)
            ws[f'{col_letter}16'] = f'={col_letter}15+{col_letter}9'
        elif sheet_config.get('rent_before_tax', False) and not sheet_config.get('dts_rts_dividend_formula', False):
            # RTS-L and RTS-ZL: =(I15+I9)*I29
            ws[f'{col_letter}16'] = f'=({col_letter}15+{col_letter}9)*{col_letter}29'
        elif sheet_config.get('dts_rts_dividend_formula', False):
            # (DTS+RTS)-L and (DTS+RTS)-ZL: Empty
            pass  # Leave empty
        else:
            # NTS-L, ITS-L, NTS-ZL, ITS-ZL: =I15*I29
            ws[f'{col_letter}16'] = f'={col_letter}15*{col_letter}29'
    
    # Row 17: NOI APPROACH
    ws['A17'] = "NOI APPROACH"
    ws['G17'] = "NOI APPROACH"
    ws['H17'] = "NOI APPROACH"
    
    # Different NOI formulas based on what was deducted before tax
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        
        if sheet_config.get('interest_before_tax', False):
            # ITS-L and ITS-ZL: Add back interest (=I12+I5+I7)
            ws[f'{col_letter}17'] = f'={col_letter}12+{col_letter}5+{col_letter}7'
        elif sheet_config.get('rent_before_tax', False) and not sheet_config.get('dts_rts_dividend_formula', False):
            # RTS-L and RTS-ZL: Add back rent (=I12+I6+I7)
            ws[f'{col_letter}17'] = f'={col_letter}12+{col_letter}6+{col_letter}7'
        elif sheet_config.get('dts_dividend_formula', False):
            # DTS-L and DTS-ZL: Add back dividend (=I12+I9+I7)
            ws[f'{col_letter}17'] = f'={col_letter}12+{col_letter}9+{col_letter}7'
        elif sheet_config.get('dts_rts_dividend_formula', False):
            # (DTS+RTS)-L and (DTS+RTS)-ZL: Add back both rent and dividend (=I12+I6+I9+I7)
            ws[f'{col_letter}17'] = f'={col_letter}12+{col_letter}6+{col_letter}9+{col_letter}7'
        else:
            # NTS-L and NTS-ZL: Standard formula (=I12+I7)
            ws[f'{col_letter}17'] = f'={col_letter}12+{col_letter}7'
    
    # Row 18: Share Capital
    ws['A18'] = "Share Capital"
    ws['H18'] = "EQUITY"
    # Set initial value I18 = 30000, then cascading: J18=I18-10.5, K18=J18-10.5, etc.
    col_letter_i = get_col_letter(9)  # Column I
    ws[f'{col_letter_i}18'] = 30000  # Initial value
    for i in range(1, n_iter):  # Start from J column (i=1)
        col_letter = get_col_letter(9 + i)
        prev_col_letter = get_col_letter(9 + i - 1)
        ws[f'{col_letter}18'] = f'={prev_col_letter}18-10.5'
    
    # Row 19: Market value of debt - KEY DIFFERENCE FOR -ZL SHEETS
    ws['A19'] = "Market value of debt"
    ws['D19'] = "Equity"
    ws['F19'] = 28000
    ws['G19'] = "TOTAL DEBT"
    ws['H19'] = "DEBT"
    
    col_letter_i = get_col_letter(9)  # Column I
    if sheet_config.get('zero_debt', False):
        # -ZL sheets: I19 = 0, then blank for J19 onwards
        ws[f'{col_letter_i}19'] = 0
        # Leave J19 onwards blank (empty)
    else:
        # -L sheets: I19 = 40000, then cascading: J19=I19-40, K19=J19-40, etc.
        ws[f'{col_letter_i}19'] = 40000  # Initial value
        for i in range(1, n_iter):  # Start from J column (i=1)
            col_letter = get_col_letter(9 + i)
            prev_col_letter = get_col_letter(9 + i - 1)
            ws[f'{col_letter}19'] = f'={prev_col_letter}19-40'
    
    # Row 20: Market value of sukuk
    ws['A20'] = "Market value of sukuk"
    ws['D20'] = "Debt"
    ws['E20'] = 0
    ws['F20'] = 2000
    ws['G20'] = "TOTAL SUKUK"
    ws['H20'] = "SUKUK"
    
    col_letter_i = get_col_letter(9)  # Column I
    if sheet_config.get('zero_debt', False):
        # -ZL sheets: I20 = 40000, then cascading: J20=I20+10.5, K20=J20+10.5, etc.
        ws[f'{col_letter_i}20'] = 40000  # Initial value
        for i in range(1, n_iter):  # Start from J column (i=1)
            col_letter = get_col_letter(9 + i)
            prev_col_letter = get_col_letter(9 + i - 1)
            ws[f'{col_letter}20'] = f'={prev_col_letter}20+10.5'
    else:
        # -L sheets: I20 = 0, then cascading: J20=I20+50.5, K20=J20+50.5, etc.
        ws[f'{col_letter_i}20'] = 0  # Initial value
        for i in range(1, n_iter):  # Start from J column (i=1)
            col_letter = get_col_letter(9 + i)
            prev_col_letter = get_col_letter(9 + i - 1)
            ws[f'{col_letter}20'] = f'={prev_col_letter}20+50.5'
    
    # Row 21: Total Assets
    ws['A21'] = "Total Assets"
    ws['D21'] = "Sukuk"
    ws['E21'] = 42000
    ws['F21'] = 2000
    ws['G21'] = "TOTAL ASSETS"
    ws['H21'] = "Total Assets"
    # Formula: =I18+I19+I20 (and similar for other columns)
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}21'] = f'={col_letter}18+{col_letter}19+{col_letter}20'
    
    # Row 22: Tax rate
    ws['A22'] = "Tax rate"
    ws['D22'] = "TAX RATE"
    ws['E22'] = 0.35
    ws['F22'] = 0.35
    ws['G22'] = "TAX RATE"
    ws['H22'] = "Tax rate"
    # Constant value across all iterations
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}22'] = 0.35
    
    # Row 23: TAX SHIELD
    ws['A23'] = "TAX SHIELD (1-Tr) or (1-35%)"
    ws['H23'] = "TAX SHIELD (1-Tr) or (1-35%)"
    # Formula: =(1-I22) (and similar for other columns)
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}23'] = f'=(1-{col_letter}22)'
    
    # Row 24: Interest rate - KEY DIFFERENCE BETWEEN SHEETS
    ws['A24'] = "Interest rate"
    ws['D24'] = "INTEREST rate that is is tax deductible"
    ws['E24'] = 0.1
    ws['G24'] = "ki"
    ws['H24'] = "Interest rate"
    
    # Set F24 based on sheet configuration
    if sheet_config.get('interest_before_tax', False):
        ws['F24'] = 0.1  # ITS-L and ITS-ZL
    else:
        ws['F24'] = 0    # All other sheets
    
    # I24 references F24
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}24'] = '=F24'
    
    # Row 25: ijarah sukuk rent rate - KEY DIFFERENCE BETWEEN SHEETS
    ws['A25'] = "ijarah sukuk rent rate"
    ws['D25'] = "RENT rate that is is tax deductible"
    ws['E25'] = 0.1
    ws['G25'] = "ks"
    ws['H25'] = "ijarah sukuk rent rate"
    
    # Set F25 based on sheet configuration
    if sheet_config.get('rent_before_tax', False):
        ws['F25'] = 0.1  # RTS-L, RTS-ZL, (DTS+RTS)-L, (DTS+RTS)-ZL
    else:
        ws['F25'] = 0    # All other sheets
    
    # I25 references F25
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}25'] = '=F25'
    
    # Row 26: ijrah depreciation benefit
    ws['A26'] = "ijrah depreciation benefit/or daily running expenses"
    ws['D26'] = "NDTS rate that is is tax deductible"
    ws['E26'] = 0.03
    ws['F26'] = 0.025
    ws['G26'] = "NDTS RATE"
    ws['H26'] = "ijrah depreciation benefit/or daily running expenses"
    # Constant value
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}26'] = 0.025
    
    # Row 27: Capital gain or loss
    ws['A27'] = "In case of purchase back asset from sukuk holder"
    ws['D27'] = "Capital gain or loss that is tax deductible"
    ws['E27'] = 0.03
    ws['F27'] = 0.03
    ws['G27'] = "CAPITAL GAIN OR LOSS"
    ws['H27'] = "In case of purchase back asset from sukuk holder"
    # Constant value
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}27'] = 0.03
    
    # Row 28: dividend rate
    ws['A28'] = "dividend rate"
    ws['D28'] = "DIVIDEND rate that is is tax deductible"
    ws['E28'] = 0.5
    ws['G28'] = "DIVIDEND RATE"
    ws['H28'] = "dividend rate"
    
    # Set F28 based on sheet type
    if sheet_name in ['RTS-L', 'DTS-L', '(DTS+RTS)-L', 'RTS-ZL', 'DTS-ZL', '(DTS+RTS)-ZL']:
        ws['F28'] = 0.5
    else:
        ws['F28'] = 0
    # Special pattern: I28=F28, J28=I28, K28=J28, L28=K28, etc.
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        if i == 0:  # I28
            ws[f'{col_letter}28'] = '=F28'
        elif i == 1:  # J28
            ws[f'{col_letter}28'] = '=I28'
        else:  # K28 onwards - reference previous column
            prev_col_letter = get_col_letter(9 + i - 1)
            ws[f'{col_letter}28'] = f'={prev_col_letter}28'
    
    # Row 29: DIVIDEND RATE
    ws['G29'] = "DIVIDEND RATE"
    # Formula: =E28 (and similar for other columns)
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}29'] = f'=E28'
    
    # Row 30: ki -sheet_name
    ws['A30'] = f"ki -{sheet_name}"
    ws['C30'] = "Ki"
    ws['G30'] = "Ki"
    ws['H30'] = f"ki -{sheet_name}"
    # Formula pattern: I30=E24, J30=I30, K30=J30, etc.
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        if i == 0:  # I30
            ws[f'{col_letter}30'] = '=E24'
        else:  # J30 onwards - reference previous cell
            prev_col_letter = get_col_letter(9 + i - 1)
            ws[f'{col_letter}30'] = f'={prev_col_letter}30'
    
    # Row 31: ks -sheet_name
    ws['A31'] = f"ks -{sheet_name}"
    ws['C31'] = "Ks"
    ws['G31'] = "Ks"
    ws['H31'] = f"ks -{sheet_name}"
    # Linear progression from 0.02 to 0.44 over 1000 iterations
    start_ks = 0.02
    end_ks = 0.44
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        # Calculate linear progression: start + (end-start) * i/(n-1)
        value = start_ks + (end_ks - start_ks) * i / (n_iter - 1)
        ws[f'{col_letter}31'] = value
    
    # Row 32: ks+g
    ws['H32'] = "ks+g"
    
    # Different formulas based on sheet type
    if sheet_name in ['RTS-L', 'RTS-ZL', 'DTS-ZL', '(DTS+RTS)-L', '(DTS+RTS)-ZL']:
        # For RTS-L, RTS-ZL, DTS-ZL, (DTS+RTS)-L, and (DTS+RTS)-ZL: I32 = I31*0.65, J32 = J31*0.65, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}32'] = f'={col_letter}31*0.65'
    
    # Row 33: ke -sheet_name
    ws['A33'] = f"ke -{sheet_name}"
    ws['C33'] = "Ke"
    ws['G33'] = "Ke"
    ws['H33'] = f"ke -{sheet_name}"
    
    # Different formulas based on sheet type
    if sheet_name in ['DTS-L', 'DTS-ZL', '(DTS+RTS)-L', '(DTS+RTS)-ZL']:
        # For DTS-L, DTS-ZL, (DTS+RTS)-L, and (DTS+RTS)-ZL: I33 = I9/I18, J33 = J9/J18, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}33'] = f'={col_letter}9/{col_letter}18'
    else:
        # Formula pattern: I33=I16/I18, J33=J16/J18, K33=K16/K18, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}33'] = f'={col_letter}16/{col_letter}18'
    
    # Row 34: ko
    ws['A34'] = "ko"
    ws['G34'] = "ko"
    ws['H34'] = "ko"
    
    # Different formulas based on sheet type
    if sheet_name in ['RTS-L', 'RTS-ZL']:
        # For RTS-L and RTS-ZL: I34 = (I30*I39)+(I36*I40)+(I33*I41), ...
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}34'] = f'=({col_letter}30*{col_letter}39)+({col_letter}36*{col_letter}40)+({col_letter}33*{col_letter}41)'
    elif sheet_name in ['DTS-L', 'DTS-ZL']:
        # For DTS-L and DTS-ZL: I34 = (I30*I39)+(I31*I40)+(I37*I41), ...
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}34'] = f'=({col_letter}30*{col_letter}39)+({col_letter}31*{col_letter}40)+({col_letter}37*{col_letter}41)'
    elif sheet_name in ['(DTS+RTS)-L', '(DTS+RTS)-ZL']:
        # For (DTS+RTS)-L and (DTS+RTS)-ZL: I34 = (I30*I39)+(I36*I40)+(I37*I41), ...
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}34'] = f'=({col_letter}30*{col_letter}39)+({col_letter}36*{col_letter}40)+({col_letter}37*{col_letter}41)'
    else:
        # Formula: =(I30*I39)+(I31*I40)+(I33*I41) (and similar for other columns)
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}34'] = f'=({col_letter}30*{col_letter}39)+({col_letter}31*{col_letter}40)+({col_letter}33*{col_letter}41)'
    
    # Row 35: Formula =I30*I23, =J30*J23, =K30*K23, etc.
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}35'] = f'={col_letter}30*{col_letter}23'
    
    # Row 36: ks with tax shield
    ws['A36'] = "ks with tax shield"
    ws['G36'] = "ks (1-Tr)"
    ws['H36'] = "ks with tax shield"
    
    # Different formulas based on sheet type
    if sheet_name in ['RTS-L', 'RTS-ZL', 'DTS-ZL', '(DTS+RTS)-L', '(DTS+RTS)-ZL']:
        # For RTS-L, RTS-ZL, DTS-ZL, (DTS+RTS)-L, and (DTS+RTS)-ZL: I36 = I32*I23, J36 = J32*J23, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}36'] = f'={col_letter}32*{col_letter}23'
    else:
        # Formula: =I31*I23, =J31*J23, =K31*K23, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}36'] = f'={col_letter}31*{col_letter}23'
    
    # Row 37: ke with tax shield
    ws['A37'] = "ke with tax shield"
    ws['G37'] = "ke (1-Tr)"
    ws['H37'] = "ke with tax shield"
    # Formula: =I33*I23, =J33*J23, =K33*K23, etc.
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}37'] = f'={col_letter}33*{col_letter}23'
    
    # Row 38: ko
    ws['G38'] = "ko"
    # Empty row - no formulas
    
    # Row 39: DEBT-L (Debt to Assets ratio)
    ws['A39'] = "DEBT-L"
    ws['G39'] = "DEBT TO ASSETS"
    ws['H39'] = "DEBT TO ASSETS"
    # Formula: =I19/I21 (and similar for other columns)
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}39'] = f'={col_letter}19/{col_letter}21'
    
    # Row 40: SUKUK-L (Sukuk to Assets ratio)
    ws['A40'] = "SUKUK-L"
    ws['G40'] = "SUKUK TO ASSETS"
    ws['H40'] = "SUKUK TO ASSETS"
    # Formula: =I20/I21 (and similar for other columns)
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}40'] = f'={col_letter}20/{col_letter}21'
    
    # Row 41: EQUITY-L (Equity to Assets ratio)
    ws['A41'] = "EQUITY-L"
    ws['G41'] = "EQUITY TO ASSETS"
    ws['H41'] = "EQUITY TO ASSETS"
    # Formula: =I18/I21 (and similar for other columns)
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}41'] = f'={col_letter}18/{col_letter}21'
    
    # Row 42: WACC -sheet_name
    ws['A42'] = f"WACC -{sheet_name}"
    ws['G42'] = "WACC"
    ws['H42'] = f"WACC -{sheet_name}"
    # Formula: =I34, =J34, =K34, etc.
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}42'] = f'={col_letter}34'
    
    # Row 43: Market Value of Firm (MVF)
    ws['G43'] = "Market Value of Firm (MVF)"
    # Formula: =I17/I42, =J17/J42, =K17/K42, etc.
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}43'] = f'={col_letter}17/{col_letter}42'
    
    # Row 44: Annual TAX SHIELD benefits of debt
    ws['A44'] = "Annual TAX SHIELD benefits of debt"
    ws['G44'] = "Annual TAX SHIELD benefits of debt"
    ws['H44'] = "Annual TAX SHIELD benefits of debt"
    # Formula: =I5*I22, =J5*J22, =K5*K22, etc.
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}44'] = f'={col_letter}5*{col_letter}22'
    
    # Row 45: Annual RENT SHIELD benefits of sukuk
    ws['A45'] = "Annual RENT SHIELD benefits of sukuk"
    ws['G45'] = "Annual RENT SHIELD benefits of sukuk"
    ws['H45'] = "Annual RENT SHIELD benefits of sukuk"
    # Formula: =I22*(I6+I7), =J22*(J6+J7), =K22*(K6+K7), etc.
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}45'] = f'={col_letter}22*({col_letter}6+{col_letter}7)'
    
    # Row 46: Annual Dividend SHIELD benefits of Equity
    ws['A46'] = "Annual Dividend SHIELD benefits of Equity"
    ws['G46'] = "Annual Dividend SHIELD benefits of Equity"
    ws['H46'] = "Annual Dividend SHIELD benefits of Equity"
    # Empty row - no formulas
    
    # Row 47: PV of TAX SHIELD benefits of debt
    ws['A47'] = "PV of TAX SHIELD benefits of debt"
    ws['G47'] = "PV of INTEREST TAX SHIELD benefits of debt"
    ws['H47'] = "PV of TAX SHIELD benefits of debt"
    # Formula: =I22*I19, =J22*J19, =K22*K19, etc.
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}47'] = f'={col_letter}22*{col_letter}19'
    
    # Row 48: PV of RENT SHIELD benefits of sukuk
    ws['A48'] = "PV of RENT SHIELD benefits of sukuk"
    ws['G48'] = "PV of RENTAL TAX SHIELD benefits of sukuk"
    ws['H48'] = "PV of RENT SHIELD benefits of sukuk"
    # Formula: =I20*I22, =J20*J22, =K20*K22, etc.
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}48'] = f'={col_letter}20*{col_letter}22'
    
    # Row 49: PV of Dividend SHIELD benefits of Equity
    ws['A49'] = "PV of Dividend SHIELD benefits of Equity"
    ws['G49'] = "PV of Dividend TAX SHIELD benefits of Equity"
    ws['H49'] = "PV of Dividend SHIELD benefits of Equity"
    
    # Different formulas based on sheet type
    if sheet_name in ['DTS-L', 'DTS-ZL', '(DTS+RTS)-L', '(DTS+RTS)-ZL']:
        # For DTS-L, DTS-ZL, (DTS+RTS)-L, and (DTS+RTS)-ZL: I49 = I9*I22/I37, J49 = J9*J22/J37, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}49'] = f'={col_letter}9*{col_letter}22/{col_letter}37'
    # For all other sheets, leave Row 49 empty (no formulas)
    
    # Row 50: PV of Bankruptcy Cost
    ws['A50'] = "PV of Bankruptcy Cost"
    ws['G50'] = "PV of Bankruptcy Cost"
    ws['H50'] = "PV of Bankruptcy Cost"
    # Empty row - no formulas
    
    # Row 51: MVF-sheet_name
    ws['A51'] = f"MVF-{sheet_name}"
    ws['G51'] = "MVF + Present Value of Tax Shield/Bc"
    ws['H51'] = f"MVF-{sheet_name}"
    
    # Different formulas based on sheet type
    if sheet_name in ['RTS-L', 'RTS-ZL']:
        # For RTS-L and RTS-ZL: I51 = I43-I47+I48, J51 = J43-J47+J48, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}51'] = f'={col_letter}43-{col_letter}47+{col_letter}48'
    elif sheet_name in ['DTS-L', 'DTS-ZL']:
        # For DTS-L and DTS-ZL: I51 = I43-I47+I49, J51 = J43-J47+J49, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}51'] = f'={col_letter}43-{col_letter}47+{col_letter}49'
    elif sheet_name in ['(DTS+RTS)-L', '(DTS+RTS)-ZL']:
        # For (DTS+RTS)-L and (DTS+RTS)-ZL: I51 = I43-I47+I48+I49, J51 = J43-J47+J48+J49, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}51'] = f'={col_letter}43-{col_letter}47+{col_letter}48+{col_letter}49'
    else:
        # Formula: =I43-I47, =J43-J47, =K43-K47, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}51'] = f'={col_letter}43-{col_letter}47'
    
    # Row 52: No. of Shares Outstanding
    ws['A52'] = "No. of Shares Outstanding"
    ws['G52'] = "NO OF SHARES OUTSTANDING"
    ws['H52'] = "No. of Shares Outstanding"
    # Formula: =I18/5, =J18/5, =K18/5, etc.
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}52'] = f'={col_letter}18/5'
    
    # Row 53: EPS-sheet_name
    ws['A53'] = f"EPS-{sheet_name}"
    ws['G53'] = "EARNING PER SHARE"
    ws['H53'] = f"EPS-{sheet_name}"
    
    # Different formulas based on sheet type
    if sheet_name in ['DTS-L', 'DTS-ZL', '(DTS+RTS)-L', '(DTS+RTS)-ZL']:
        # For DTS-L, DTS-ZL, (DTS+RTS)-L, and (DTS+RTS)-ZL: I53 = I9/I52, J53 = J9/J52, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}53'] = f'={col_letter}9/{col_letter}52'
    else:
        # Formula: =I16/I52, =J16/J52, =K16/K52, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}53'] = f'={col_letter}16/{col_letter}52'
    
    # Create the additional calculation rows (74-76, 114-122) with dynamic formulas
    # Row 74: MVF (Scaled Value)
    ws['A74'] = "MVF (Scaled Value)"
    ws['G74'] = "MVF (Scaled Value)"
    ws['H74'] = "MVF (Scaled Value)"
    
    # Different formulas based on sheet type
    if sheet_name in ['RTS-L', 'RTS-ZL', 'DTS-L', 'DTS-ZL', '(DTS+RTS)-L', '(DTS+RTS)-ZL']:
        # For RTS-L, RTS-ZL, DTS-L, DTS-ZL, (DTS+RTS)-L, and (DTS+RTS)-ZL: I74 = I51/32000, J74 = J51/32000, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}74'] = f'={col_letter}51/32000'
    else:
        # Formula: =I43/32000, =J43/32000, etc.
        for i in range(n_iter):
            col_letter = get_col_letter(9 + i)
            ws[f'{col_letter}74'] = f'={col_letter}43/32000'
    
    # Row 75: Tax Contribution Scalled
    ws['A75'] = "Tax Contribution Scalled"
    ws['G75'] = "Tax Contribution Scalled"
    ws['H75'] = "Tax Contribution Scalled"
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}75'] = f'={col_letter}11/1700'
    
    # Row 76: T.C
    ws['A76'] = "T.C"
    ws['G76'] = "T.C"
    ws['H76'] = "T.C"
    
    # Different formulas based on sheet type
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        
        if sheet_name in ['ITS-L']:
            # ITS-L: I76=I31+I33+I35, J76=J31+J33+J35, ...
            ws[f'{col_letter}76'] = f'={col_letter}31+{col_letter}33+{col_letter}35'
        elif sheet_name in ['RTS-L', 'RTS-ZL']:
            # RTS-L and RTS-ZL: I76=I30+I33+I36, J76=J30+J33+J36, ...
            ws[f'{col_letter}76'] = f'={col_letter}30+{col_letter}33+{col_letter}36'
        elif sheet_name in ['DTS-L', 'DTS-ZL']:
            # DTS-L and DTS-ZL: I76=I30+I31+I37, J76=J30+J31+J37, ...
            ws[f'{col_letter}76'] = f'={col_letter}30+{col_letter}31+{col_letter}37'
        elif sheet_name in ['(DTS+RTS)-L', '(DTS+RTS)-ZL']:
            # (DTS+RTS)-L and (DTS+RTS)-ZL: I76=I30+I36+I37, J76=J30+J36+J37, ...
            ws[f'{col_letter}76'] = f'={col_letter}30+{col_letter}36+{col_letter}37'
        else:
            # NTS-L, NTS-ZL, ITS-ZL: I76=I30+I31+I33, J76=J30+J31+J33, ...
            ws[f'{col_letter}76'] = f'={col_letter}30+{col_letter}31+{col_letter}33'
    
    # Row 114: Share Capital %
    ws['A114'] = "Share Capital"
    ws['G114'] = "SHARE CAPITAL"
    ws['H114'] = "Equity %"
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}114'] = f'={col_letter}18/{col_letter}21*100'
    
    # Row 115: Market value of debt %
    ws['A115'] = "Market value of debt"
    ws['G115'] = "TOTAL DEBT"
    ws['H115'] = "Debt %"
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}115'] = f'={col_letter}19/{col_letter}21*100'
    
    # Row 116: Market value of sukuk %
    ws['A116'] = "Market value of sukuk"
    ws['G116'] = "TOTAL SUKUK"
    ws['H116'] = "Sukuk %"
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}116'] = f'={col_letter}20/{col_letter}21*100'
    
    # Row 117: Total Assets %
    ws['A117'] = "Total Assets"
    ws['G117'] = "TOTAL ASSETS"
    ws['H117'] = "Total Assets %"
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}117'] = f'={col_letter}114+{col_letter}115+{col_letter}116'
    
    # Row 119: Share Capital % (duplicate)
    ws['A119'] = "Share Capital"
    ws['G119'] = "SHARE CAPITAL"
    ws['H119'] = "Equity %"
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}119'] = f'={col_letter}114'
    
    # Row 120: Market value of debt % (duplicate)
    ws['A120'] = "Market value of debt"
    ws['G120'] = "TOTAL DEBT"
    ws['H120'] = "Debt %"
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}120'] = f'={col_letter}115'
    
    # Row 121: Market value of sukuk % (duplicate)
    ws['A121'] = "Market value of sukuk"
    ws['G121'] = "TOTAL SUKUK"
    ws['H121'] = "Sukuk %"
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}121'] = f'={col_letter}116'
    
    # Row 122: Total Assets % (duplicate)
    ws['A122'] = "Total Assets"
    ws['G122'] = "TOTAL ASSETS"
    ws['H122'] = "Total Assets %"
    for i in range(n_iter):
        col_letter = get_col_letter(9 + i)
        ws[f'{col_letter}122'] = f'={col_letter}117'
    
    # Now add row grouping and collapse the empty ranges to match original sheet format
    # Group and collapse rows 54-73 (empty range)
    for row in range(54, 74):
        ws.row_dimensions[row].outline_level = 1
        ws.row_dimensions[row].hidden = True
    
    # Group and collapse rows 78-113 (empty range)  
    for row in range(78, 114):
        ws.row_dimensions[row].outline_level = 1
        ws.row_dimensions[row].hidden = True
    
    # Also hide row 77 (empty) and row 118 (empty)
    ws.row_dimensions[77].hidden = True
    ws.row_dimensions[118].hidden = True


# Define configurations for each sheet
sheet_configs = {
    'NTS-L': {
        'interest_before_tax': False,
        'rent_before_tax': False,
        'zero_debt': False,
        'no_dividend_before_tax': False,
        'dts_dividend_formula': False,
        'dts_rts_dividend_formula': False,
        'rent_uses_31': False
    },
    'ITS-L': {
        'interest_before_tax': True,
        'rent_before_tax': False,
        'zero_debt': False,
        'no_dividend_before_tax': False,
        'dts_dividend_formula': False,
        'dts_rts_dividend_formula': False,
        'rent_uses_31': False
    },
    'RTS-L': {
        'interest_before_tax': False,
        'rent_before_tax': True,
        'zero_debt': False,
        'no_dividend_before_tax': True,
        'dts_dividend_formula': False,
        'dts_rts_dividend_formula': False,
        'rent_uses_31': True
    },
    'DTS-L': {
        'interest_before_tax': False,
        'rent_before_tax': False,
        'zero_debt': False,
        'no_dividend_before_tax': False,
        'dts_dividend_formula': True,
        'dts_rts_dividend_formula': False,
        'rent_uses_31': False
    },
    '(DTS+RTS)-L': {
        'interest_before_tax': False,
        'rent_before_tax': True,
        'zero_debt': False,
        'no_dividend_before_tax': False,
        'dts_dividend_formula': False,
        'dts_rts_dividend_formula': True,
        'rent_uses_31': True
    },
    'NTS-ZL': {
        'interest_before_tax': False,
        'rent_before_tax': False,
        'zero_debt': True,
        'no_dividend_before_tax': False,
        'dts_dividend_formula': False,
        'dts_rts_dividend_formula': False,
        'rent_uses_31': False
    },
    'ITS-ZL': {
        'interest_before_tax': True,
        'rent_before_tax': False,
        'zero_debt': True,
        'no_dividend_before_tax': False,
        'dts_dividend_formula': False,
        'dts_rts_dividend_formula': False,
        'rent_uses_31': False
    },
    'RTS-ZL': {
        'interest_before_tax': False,
        'rent_before_tax': True,
        'zero_debt': True,
        'no_dividend_before_tax': True,
        'dts_dividend_formula': False,
        'dts_rts_dividend_formula': False,
        'rent_uses_31': True
    },
    'DTS-ZL': {
        'interest_before_tax': False,
        'rent_before_tax': False,
        'zero_debt': True,
        'no_dividend_before_tax': False,
        'dts_dividend_formula': True,
        'dts_rts_dividend_formula': False,
        'rent_uses_31': False
    },
    '(DTS+RTS)-ZL': {
        'interest_before_tax': False,
        'rent_before_tax': True,
        'zero_debt': True,
        'no_dividend_before_tax': False,
        'dts_dividend_formula': False,
        'dts_rts_dividend_formula': True,
        'rent_uses_31': True
    }
}

# Create a new workbook
wb = Workbook()

# Create all sheets
for sheet_name, config in sheet_configs.items():
    print(f"Creating sheet: {sheet_name}")
    create_sheet(wb, sheet_name, config)

# Save the workbook
output_file = 'sukuk_analysis_1000_iterations.xlsx'
wb.save(output_file)

print(f"\nExcel file with all sheets and 1000 iterations saved as {output_file}")
print(f"Total iterations: {n_iter}")
print("Created sheets:", list(sheet_configs.keys()))
print("\nSheet configurations:")
for sheet_name, config in sheet_configs.items():
    print(f"  {sheet_name}: {config}")